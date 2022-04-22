import time
import pandas as pd
import matplotlib.pylab as plt
import pulp as p
#import winsound
import os
import openpyxl
from random import randint

def Adjacent_Swap(Node_Sequence):
    Upto_Len=len(Node_Sequence)-2
    temp=randint(1, Upto_Len)
    Node_Sequence[temp]-=Node_Sequence[temp+1]
    Node_Sequence[temp+1]+=Node_Sequence[temp]
    Node_Sequence[temp]=Node_Sequence[temp+1]-Node_Sequence[temp]

def General_Swap(Node_Sequence):
    Upto_Len=len(Node_Sequence)-1
    temp1=randint(1, Upto_Len)
    temp2=randint(1, Upto_Len)
    if temp1==temp2:
        General_Swap(Node_Sequence)
    temp=Node_Sequence[temp1]
    Node_Sequence[temp1]=Node_Sequence[temp2]
    Node_Sequence[temp2]=temp

def Single_Insertion(Node_Sequence):
    Upto_Len=len(Node_Sequence)-1-2
    temp1=randint(1, Upto_Len)
    temp2=Node_Sequence[temp1]
    del Node_Sequence[temp1]
    temp3=randint(temp1+2,Upto_Len+2)
    Node_Sequence.insert(temp3,temp2)

def Reversal(Node_Sequence):
    Upto_Len=len(Node_Sequence)
    start=randint(1,Upto_Len-3)
    finish=randint(start+2,Upto_Len-1)
    Node_Sequence=Node_Sequence[:start:+1]+Node_Sequence[finish:start-1:-1]+Node_Sequence[finish+1::+1]

def Decoding_Mechanism(Node_Sequence):
    if Node_Sequence[0]!=0:
        print("EXCEPTION: Please Check Code")
    # Node Sequence should be an array starting from 0th Node and end at the last Node before the Depot, example 0,1,8,6,7
    set_of_routes={}
    Upto_Len=len(Node_Sequence)-1
    for i in range(Upto_Len):
        for k in Vehicle_Types:
            DynamicCapacityLeft=[VQ[k]]
            Edge=VC[k]+VS[k]*C[0,Node_Sequence[i+1],k]
            for j in range(i+1,Upto_Len+1):
                CapacityCheck=0
                #print(DynamicCapacityLeft[0])
                DynamicCapacityLeft[0]=DynamicCapacityLeft[0]-d[Node_Sequence[j]]
                DynamicCapacityLeft.append(d[Node_Sequence[j]]-pp[Node_Sequence[j]])
                for m in DynamicCapacityLeft:
                    CapacityCheck+=m
                    if CapacityCheck<0:
                        break
                if CapacityCheck<0:
                    break
                #key=(Node_Sequence[i],Node_Sequence[j],k)
                key=(i,j,k)
                if j>i+1:
                    Edge+=VS[k]*C[Node_Sequence[j-1],Node_Sequence[j],k]
                set_of_routes[key]=Edge+VS[k]*C[Node_Sequence[j],0,k]
    #print(set_of_routes)
    another_set_of_routes=set_of_routes.copy()

    # Applying Dijktra's Algorithm / Dynamic Programming to find the shortest path among the above routes
    Min_Cost_Array=[]
    for i in range(Upto_Len+1):
        # This Dijktra / Dynamic Programming Array will have (Minimum Cost,Next Node,Vehicle Type) as its value
        # So the last Node will have (0,0,0) indicating 
        Min_Cost_Array.append((0,0,0))
    for i in range(Upto_Len-1,-1,-1):
        minim=999999999 # This should be a Very Large Number
        for j in another_set_of_routes:
            delete_bin=set()
            if j[0]==i:
                compare=another_set_of_routes[j]+Min_Cost_Array[j[1]][0]
                #print("This is compare",compare)
                #print("this is min",min)
                if compare<minim:
                    minim=compare
                    Min_Cost_Array[i]=(minim,j[1],j[2])
                delete_bin.add(j) # Removing the used Edge from the subsequent search space after the loop is done
        for j in delete_bin:
            del another_set_of_routes[j]
    # The Min_Cost_Array refers to the indexes of the Node_Sequences

    message="Routes:- \n "
    interval_start=0
    inter_startup=0
    Next_Node=Min_Cost_Array[0][1]
    while Next_Node!=0:
        message+=" Vehicle Type "+str(Min_Cost_Array[inter_startup][2])+" : \t "
        if interval_start!=0:
            message+=str(0)+" --> "
        for i in range(interval_start,Next_Node+1):
            message+=str(Node_Sequence[i])+" --> "
        message+="0 \n "
        inter_startup=Next_Node
        interval_start=Next_Node+1
        Next_Node=Min_Cost_Array[Next_Node][1]
    
    start=0
    solution=set()
    Next_Node=Min_Cost_Array[0][1]
    while Next_Node!=0:
        solution.add((0,Node_Sequence[start+1],Min_Cost_Array[start][2]))
        for i in range(start+1,Next_Node):
            solution.add((Node_Sequence[i],Node_Sequence[i+1],Min_Cost_Array[start][2]))
        solution.add((Node_Sequence[Next_Node],0,Min_Cost_Array[start][2]))
        start=Next_Node
        Next_Node=Min_Cost_Array[Next_Node][1]

    return (Min_Cost_Array[0][0],message,solution)


directory_name="Routes, Images and Values of Decision Variables"
main_dir = directory_name
os.mkdir(main_dir)

# Call a Workbook() function of openpyxl to create a new blank Workbook object
wb = openpyxl.Workbook()
# Get workbook active sheet from the active attribute
sheet = wb.active
row_number_on_Excel_Table=1
cell = sheet.cell(row = row_number_on_Excel_Table, column = 1)
cell.value = "Upto Node Number considered from the 0th Node which is the Vehicle_Depot, Warehouse, as well as the final location for transporting evacuees (All Vehicles have been considered)"
cell = sheet.cell(row = row_number_on_Excel_Table, column = 2)
cell.value = "Objective Value as Obtained from the Heuristic"
cell = sheet.cell(row = row_number_on_Excel_Table, column = 3)
cell.value = "Time Taken for the Heuristic to reach this solution"
cell = sheet.cell(row = row_number_on_Excel_Table, column = 4)
cell.value = "Solution obtained from PuLP in this same time"
wb.save("Table.xlsx")

# Get the Input
Nodes=pd.read_excel("Input Data.xlsx","Locations & Delivery-PickUp",index_col=0)
range_of_number_of_nodes=Nodes.shape[0]

Vehicles=pd.read_excel("Input Data.xlsx","Vehicle Specifications",index_col=0)
num_of_Vehicle_types=Vehicles.shape[0]

#for upto_Node_number in range(1,range_of_number_of_nodes):
for upto_Node_number in range(9,19):
    row_number_on_Excel_Table+=1
    cell = sheet.cell(row = row_number_on_Excel_Table, column = 1)
    cell.value = upto_Node_number
    wb.save("Table.xlsx")

    main_dir = directory_name+"/"+str(upto_Node_number)
    os.mkdir(main_dir)
    
    #Sets Used
    Relief_Centres=set()
    Depot_and_Relief_Centres=set()

    counter=0
    for i, row in Nodes.iterrows():
        if i!=0: # PLEASE ENSURE 0 IS DEFINED AS THE DEPOT. ALSO ENSURE THE SOLO DEPOT HAS NODE NUMBER 0.
            Relief_Centres.add(i)
        Depot_and_Relief_Centres.add(i)
        counter+=1
        if counter>upto_Node_number:
            break

    counter=0
    Vehicle_Types=set()
    VN={}
    VQ={}
    VS={}
    VC={}
    for i, row in Vehicles.iterrows():
        Vehicle_Types.add(i)
        VN[i]=row["VN"]
        VQ[i]=row["VQ"]
        VS[i]=row["VS"]
        VC[i]=row["VC"]
        counter+=1
        if counter>=num_of_Vehicle_types:
            break

    pp={}
    d={}
    for i, row in Nodes.iterrows():
        pp[i]=row["PickUp"]
        d[i]=row["Delivery"]
        if i>=upto_Node_number:
            break

    #Creating the Distance Matrices
    C={} #This is the COST/DISTANCE Matrix

    #For checking the exact solution of the paper here all the distances are considered Euclidean
    for k in Vehicle_Types:
        temp = pd.read_excel("Input Data.xlsx","Calculating Random Distances")
        for i,row in temp.iterrows():
            if row["Origin Node"] in Depot_and_Relief_Centres:
                if (row["Destination Node"] in Depot_and_Relief_Centres) and (row["Destination Node"]!=row["Origin Node"]):
                    key=(int(row["Origin Node"]),int(row["Destination Node"]),int(k))
                    value=float(row["Euclidean Distance"])
                    C[key]=value
    """
    for k in Vehicle_Types:
        temp = pd.read_excel("Input Data.xlsx",str(k))
        for i,row in temp.iterrows():
            if row["Origin Node"] in Depot_and_Relief_Centres:
                if (row["Destination Node"] in Depot_and_Relief_Centres) and (row["Destination Node"]!=row["Origin Node"]):
                    key=(int(row["Origin Node"]),int(row["Destination Node"]),int(k))
                    value=float(row["Distance"])
                    C[key]=value
    """

    # Heuristic starts here
    """
    set_of_routes={}
    for i in range(upto_Node_number):
        for k in Vehicle_Types:
            DynamicCapacityLeft=[VQ[k]]
            Edge=VC[k]+VS[k]*C[0,i+1,k]
            for j in range(i+1,upto_Node_number+1):
                CapacityCheck=0
                #print(DynamicCapacityLeft[0])
                DynamicCapacityLeft[0]=DynamicCapacityLeft[0]-d[j]
                DynamicCapacityLeft.append(d[j]-pp[j])
                for m in DynamicCapacityLeft:
                    CapacityCheck+=m
                    if CapacityCheck<0:
                        break
                if CapacityCheck<0:
                    break
                key=(i,j,k)
                if j>i+1:
                    Edge+=VS[k]*C[j-1,j,k]
                set_of_routes[key]=Edge+VS[k]*C[j,0,k]
    #print(set_of_routes)
    """
    main_dir=main_dir+"/"

    textfile = open(main_dir+"Vehicle Routes from HLS Heuristic of Paper.txt","w")
    Node_Sequence=[]
    for i in range(upto_Node_number+1):
        Node_Sequence.append(i)
    Tabu_List=[] # A set would have been better but the ingredients of the return statement of the function Decoding Mechanism contains other sets and Python does not allow sets within sets...
    N=3 # Neighbourhood Structure Number
    Ci=1
    ii=1
    x=Decoding_Mechanism(Node_Sequence)
    Tabu_List.append(x)
    x_b=x
    age=0
    f=0
    a1=x_b[0]/x[0]
    a2=Ci/ii
    t=1+a1*a2
    #f_iter=int(input("Enter the Number of iterations required"))
    f_iter=123
    start_time=time.time()
    while f<f_iter:
        ppass=0
        while ppass==0:
            Adjacent_Swap(Node_Sequence)
            General_Swap(Node_Sequence)
            Single_Insertion(Node_Sequence)
            Reversal(Node_Sequence)
            x_dash=Decoding_Mechanism(Node_Sequence)
            counter=1
            for inside_Tabu in Tabu_List:
                if inside_Tabu[2]==x_dash[2]:
                    counter=0
                    break
            if counter==1:
                ppass=1
        if x_dash[0]<=t*x_b[0]:
            x=x_dash
            ii+=1
            age=0
            Tabu_List.append(x)
            if x[0]<x_b[0]:
                f=0
                x_b=x
                Ci+=1
                for g in x_b:
                    textfile.write(str("\n"))
                    print("\n")
                    textfile.write(str(g))
                    print(g)
                textfile.write(str("\n \n"))
                print("\n \n")
            else:
                f+=1
            a1=x_b[0]/x[0]
            a2=Ci/ii
            t=1+a1*a2
        else:
            age+=1
            if age==N:
                age=0
                t=t+a1*a2

    end_time=time.time()
    delta_T=end_time-start_time
    print("The final output was obtained after "+str(delta_T)+" seconds.")
    textfile.write("The final output was obtained after "+str(delta_T)+" seconds.")
    textfile.close()

    cell = sheet.cell(row = row_number_on_Excel_Table, column = 2)
    cell.value = x_b[0]
    cell = sheet.cell(row = row_number_on_Excel_Table, column = 3)
    cell.value = delta_T
    wb.save("Table.xlsx")



    
    # Set the problem
    prob=p.LpProblem("Heterogenenous_single_Depot_mVRPSDP",p.LpMinimize)

    # Decision Variables
    # Iff Arc joining i & j is included within the solution for the Layer k
    x=p.LpVariable.dicts('x',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j),cat='Binary')

    # Amount of collected load across Arc(i,j) by a Vehicle in Layer k
    y=p.LpVariable.dicts('y',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j),lowBound=0)

    # Amount of delivery load across Arc(i,j) done by a Vehicle in Layer k
    z=p.LpVariable.dicts('z',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j),lowBound=0)

    # Set Objective Function (Point 2)
    prob+=p.lpSum(x[i,j,k]*C[i,j,k]*VS[k] for k in Vehicle_Types for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j)+p.lpSum(x[0,j,k]*VC[k] for k in Vehicle_Types for j in Relief_Centres)

    #Ensuring at most a single vehicle caters to a Relief Center (Point 3 a)
    for i in Relief_Centres:
        prob+=p.lpSum(x[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j)<=1

    # Ensuring equal number of Incoming and Outgoing paths are available from all Nodes (Point 3 b)
    for i in Depot_and_Relief_Centres:
        for k in Vehicle_Types:
            prob+=p.lpSum(x[i,j,k]-x[j,i,k] for j in Depot_and_Relief_Centres  if i!=j)==0

    # Ensuring at most VN outgoing paths are available at the Depot since there are VN[k] vehicle for each Vehicle Type (Point 3 c)
    for k in Vehicle_Types:
        prob+=p.lpSum(x[0,j,k] for j in Relief_Centres)<=VN[k]

    """Flow Limitation Constraints"""

    for k in Vehicle_Types:
        for j in Relief_Centres:
            prob+=y[0,j,k]==0   #Ensuring initial PickUp from Nodes is 0 (Point 3 d i)

    for k in Vehicle_Types:
        for i in Relief_Centres:
            prob+=z[i,0,k]==0   #Ensuring final Delivery to Nodes is 0 (Point 3 d ii)

    #Ensuring the PickUp constraints are satisfied (Point 3 e i)
    for i in Relief_Centres:
        prob+=p.lpSum(y[i,j,k]-y[j,i,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j)==pp[i]

    #Ensuring the Delivery constraints are satisfied (Point 3 e ii)
    for i in Relief_Centres:
        prob+=p.lpSum(z[j,i,k]-z[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j)==d[i]

    """Constraining the Sum of Flows to and from the Origin/Depot/Warehouse/NDRF_BASE"""
    #(Point 3 f i)
    prob+=p.lpSum(y[i,0,k] for i in Relief_Centres for k in Vehicle_Types)==p.lpSum(pp[i] for i in Relief_Centres)   #Ensuring sum of all PickUp Flow Variables to the Origin [0th Node] is equal to the total PickUps of all Nodes
    #(Point 3 f ii)
    prob+=p.lpSum(z[0,i,k] for i in Relief_Centres for k in Vehicle_Types)==p.lpSum(d[i] for i in Relief_Centres)   #Ensuring sum of all Delivery Flow Variables from the Origin [0th Node] is equal to the total Demand of all Nodes

    # Ensuring the vehicle capacity is never exceeded (Point 3 g)
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in Vehicle_Types:
                if i!=j:
                    prob+=y[i,j,k]+z[i,j,k]<=VQ[k]*x[i,j,k]

    # Solve the Problem using default CBC
    #start_time=time.time()
    #status=prob.solve(p.PULP_CBC_CMD(maxSeconds=300, msg=1, gapRel=0))
    status=prob.solve(p.PULP_CBC_CMD(timeLimit=delta_T))
    #status=prob.solve()
    #end_time=time.time()

    #winsound.Beep(555-19*upto_Node_number, 888+19*upto_Node_number) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds
    print("This is the status:- ", p.LpStatus[prob.status])
    objec_val=p.value(prob.objective)
    
    '''
    # Draw the optimal routes Layerwise
    for k in Vehicle_Types:
        plt.figure(figsize=(9,9))
        for i, row in Nodes.iterrows():
            if i>upto_Node_number:
                break
            if i==0:
                plt.scatter(row["Longitude"],row["Latitude"], c='r',marker='s')
                plt.text( row["Longitude"] + 0.33, row["Latitude"] + 0.33, "Depot")
            else:
                plt.scatter(row["Longitude"], row["Latitude"], c='black')
                plt.text(row["Longitude"] + 0.33, row["Latitude"] + 0.33, i)
        plt.title('mVRPSDC Tours for Vehicles of Type '+str(k)+" on the corresponding layer "+str(k))
        plt.ylabel("Latitude")
        plt.xlabel("Longitude")

        max=0   # Finding the maximum utilised vehicle capacity
        routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j if p.value(x[i,j,k])==1]
        #routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j and p.value(x[i,j,k])==1]
        arrowprops = dict(arrowstyle='->', connectionstyle='arc3', edgecolor='blue')
        for i, j in routes:
            utilized_capacity=y[i,j,k].varValue+z[i,j,k].varValue
            if utilized_capacity>max:
                max=utilized_capacity
            plt.annotate('', xy=[Nodes.iloc[j]["Longitude"], Nodes.iloc[j]["Latitude"]], xytext=[Nodes.iloc[i]["Longitude"], Nodes.iloc[i]["Latitude"]], arrowprops=arrowprops)    
            #plt.text((Nodes.iloc[i]["Longitude"]+Nodes.iloc[j]["Longitude"])/2, (Nodes.iloc[i]["Latitude"]+Nodes.iloc[j]["Latitude"])/2, f'{utilized_capacity}',fontweight="bold")

        print("The maximum vehicle capacity utilised ever in any tour in layer ",k," is: ",max," out of the total available",VQ[k])
        
        used_vehicles=0 # Finding the maximum number of vehicles being used
        for j in Relief_Centres:
            used_vehicles=p.value(x[0,j,k])+used_vehicles
        print("The maximum numbers of vehicles used is: ",used_vehicles," out of total available ",VN[k])
        name="Vehicles_ "+str(used_vehicles)+"--"+str(VN[k])+" and Capacity_ "+str(max)+"--"+str(VQ[k])+" with Objective Value_ "+str(objec_val)+" & Solver Time is_ "+str(end_time-start_time)+"seconds.png"
        main_dir_for_Image=main_dir+"{}"
        plt.savefig(main_dir_for_Image.format(name))
    '''
    # Writing the Routes in a Text File
    textfile = open(main_dir+"Vehicle Routes as per PuLP within equal time.txt","w")
    for k in Vehicle_Types:
        counter=0
        for j in Relief_Centres:
            if p.value(x[0,j,k])==1:
                counter+=1
                start_node=j
                textfile.write("Vehicle Type: "+str(k)+",\t Vehicle Number: "+str(counter)+", \t Route=\t 0")
                while start_node!=0:
                    textfile.write(" --> "+str(start_node))
                    for i in Depot_and_Relief_Centres:
                        if  start_node!=i and p.value(x[start_node,i,k])==1:
                            start_node=i
                            break
                if start_node==0:
                    textfile.write(" --> "+str(start_node)+"\n")
    textfile.close()
    '''
    # Call a Workbook() function of openpyxl to create a new blank Workbook object
    wb_individual = openpyxl.Workbook()
    # Get workbook active sheet from the active attribute
    sheet_individual = wb_individual.active
    row_number_on_Individual_Sheet=1
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
    cell.value = "From Node i"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
    cell.value = "To Node j"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
    cell.value = "Vehicle Type k"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
    cell.value = "x_ijk indicating whether the Arc is selected"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
    cell.value = "y_ijk indicating the amount of Pickup"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
    cell.value = "z_ijk indicating the amount of Delivery"
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in Vehicle_Types:
                if i!=j:
                    row_number_on_Individual_Sheet+=1
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
                    cell.value = i
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
                    cell.value = j
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
                    cell.value = k
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
                    cell.value = p.value(x[i,j,k])
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
                    cell.value = p.value(y[i,j,k])
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
                    cell.value = p.value(z[i,j,k])
    wb_individual.save(str(main_dir)+"Solution Details for upto Node Number "+str(upto_Node_number)+".xlsx")
    '''
    cell = sheet.cell(row = row_number_on_Excel_Table, column = 4)
    cell.value = objec_val
    wb.save("Table.xlsx")